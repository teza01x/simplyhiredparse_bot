from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from config import excel_name

def create_new_excel():
    workbook = Workbook()
    sheet = workbook.active

    sheet.append(['Platform', 'Job Link', 'Company Name', 'Job Title', 'Job Location', 'Posted Time (hours ago)', 'Statut (Post On Discord, Post On Telegram, To Post)', 'Date added'])

    workbook.save("job_offers.xlsx")


def sort_records():
    workbook = load_workbook(excel_name)
    sheet = workbook.active

    # Получаем все записи из листа
    records = list(sheet.iter_rows(values_only=True))

    # Сортируем записи по значению в колонке "Posted Time (hours ago)"
    sorted_records = sorted(records[1:], key=lambda x: x[5])

    # Обновляем лист с отсортированными записями
    for i, record in enumerate(sorted_records, start=1):
        # Используем функцию zip и оператор * для привязки значений в строке
        values = zip(range(1, len(record)+1), record)
        for j, value in values:
            sheet.cell(row=i+1, column=j, value=value)

    workbook.save(excel_name)

def remove_duplicate_records():
    workbook = load_workbook(excel_name)
    sheet = workbook.active

    link_values = {}


    for row in sheet.iter_rows(min_row=2, values_only=True):
        link = row[1]
        value = int(row[5])


        if link in link_values and value < link_values[link]:
            link_values[link] = value

        else:
            link_values[link] = value


    new_records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        link = row[1]
        value = int(row[5])


        if value == link_values[link]:
            new_records.append(row)


    sheet.delete_rows(2, sheet.max_row)


    for record in new_records:
        sheet.append(record)

    workbook.save(excel_name)


def delete_old_posts():
    workbook = load_workbook(excel_name)
    sheet = workbook.active


    now = datetime.now()
    current_date = int(now.strftime("%d"))


    rows_to_delete = []
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if int(row[7].value) < current_date:
            rows_to_delete.append(row)

    for row in rows_to_delete:
        sheet.delete_rows(row[0].row)

    workbook.save(excel_name)


def sort_for_uniq():
    workbook = load_workbook(excel_name)

    sheet = workbook.active

    unique_records = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row not in unique_records:
            unique_records.append(row)

    sheet.delete_rows(2, sheet.max_row)

    for row in unique_records:
        sheet.append(row)

    workbook.save(excel_name)


def add_new_data(data):
    workbook = load_workbook(excel_name)

    sheet = workbook.active

    for i in data:
        sheet.append(i)

    try:
        workbook.save(excel_name)
        sort_for_uniq()
        delete_old_posts()
        remove_duplicate_records()
        sort_records()
    except Exception as e:
        print("Error {}".format(e))

